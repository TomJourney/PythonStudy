import openpyxl
from pathlib import Path

# 使用openpyxl模块读取excel文件
wb = openpyxl.load_workbook(Path.cwd() / '1301-工作簿.xlsx')
print(type(wb))
# <class 'openpyxl.workbook.workbook.Workbook'>

print(wb.sheetnames)
# ['Sheet1', 'Sheet2', 'Sheet3']

sheet1 = wb['Sheet1']
print(sheet1['A1'].value)
# id

# 打印行列坐标及值
cell = sheet1['A1']
print(f'[%s][%s] = %s' % (cell.row, cell.column, cell.value))
# [1][1] = id

# 通过数字定位行列坐标(从1开始计数)
for i in range(1, 5):
    for j in range(1, 4):
        cell = sheet1.cell(i, j)
        print(f'[%s][%s] = %s' % (cell.row, cell.column, cell.value), end='  ')
    print()
# [1][1] = id  [1][2] = name  [1][3] = addr
# [2][1] = 1  [2][2] = 张三01  [2][3] = 成都01
# [3][1] = 2  [3][2] = 张三02  [3][3] = 成都02
# [4][1] = 3  [4][2] = 张三03  [4][3] = 成都03

# 13.3.5 从表中获取行和列
wb = openpyxl.load_workbook(Path.cwd() / '1301-工作簿.xlsx')
sheet1 = wb['Sheet1']
sheetTuple = tuple(sheet1['A1':'A3'])
print(sheetTuple)
# ((<Cell 'Sheet1'.A1>,), (<Cell 'Sheet1'.A2>,), (<Cell 'Sheet1'.A3>,))

# 获取行与列的单元格
print("====== 从表中获取行和列 ======")
for rowOfCellObjects in sheet1:
    for cell in rowOfCellObjects:
        print(cell.coordinate, cell.value, end=' ')
    print()
# ====== 从表中获取行和列 ======
# A1 id B1 name C1 addr
# A2 1 B2 张三01 C2 成都01
# A3 2 B3 张三02 C3 成都02


# 获取整行
print("\n======= 获取整行 =======")
rowList = list(sheet1.rows)
# rowList = [(<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>)
# , (<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.C2>)
# , (<Cell 'Sheet1'.A3>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.C3>)
# , (<Cell 'Sheet1'.A4>, <Cell 'Sheet1'.B4>, <Cell 'Sheet1'.C4>)]
print("rowList = " + str(rowList))
for rowObject in rowList:
    for cellObject in rowObject:
        print(f'%s = %s ' % (cellObject.coordinate, cellObject.value), end=' ')
    print()
# A1 = id  B1 = name  C1 = addr
# A2 = 1  B2 = 张三01  C2 = 成都01
# A3 = 2  B3 = 张三02  C3 = 成都02
# A4 = 3  B4 = 张三03  C4 = 成都03

