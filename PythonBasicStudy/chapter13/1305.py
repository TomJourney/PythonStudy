from pathlib import Path

import openpyxl

wb = openpyxl.Workbook()
# 使用Workbook对象的active属性获取工作簿的活动表
sheet01 = wb.active
print(sheet01.title)  # Sheet
sheet01.title = 'PythonWriteSheet01'
print(wb.sheetnames)  # ['PythonWriteSheet01']
# 保存到xlsx文档
wb.save(Path.cwd() / '1305-01工作簿.xlsx')
wb.close()

## 创建和删除工作表
wb = openpyxl.Workbook()
# 创建工作表
wb.create_sheet("PythonWriteSheet10")
print(wb.sheetnames)
# ['Sheet', 'PythonWriteSheet02']

# 创建工作表
wb.create_sheet(index=0, title="PythonWriteSheet00")
wb.create_sheet(index=1, title="PythonWriteSheet01")
wb.create_sheet(index=2, title="PythonWriteSheet02")
wb.create_sheet(index=3, title="PythonWriteSheet03")
print(wb.sheetnames)
# ['PythonWriteSheet00', 'PythonWriteSheet01', 'PythonWriteSheet02', 'PythonWriteSheet03', 'Sheet', 'PythonWriteSheet10']

# 删除工作表
del wb['PythonWriteSheet10']
print(wb.sheetnames)
# ['PythonWriteSheet00', 'PythonWriteSheet01', 'PythonWriteSheet02', 'PythonWriteSheet03', 'Sheet']

# 保存excel文件到磁盘
wb.save(Path.cwd() / '1305-02工作簿.xlsx')
wb.close()


# 把值写入单元格
wb = openpyxl.Workbook()
# 创建工作表
wb.create_sheet("PythonWriteSheetA01")
sheetA01 = wb["PythonWriteSheetA01"]
sheetA01.cell(1, 1, "值1-1")
sheetA01.cell(1, 2, "值1-2")
sheetA01.cell(2, 1, "值2-1")
sheetA01.cell(2, 2, "值2-2")
wb.save(Path.cwd() / '1305-工作簿A01.xlsx')
wb.close()